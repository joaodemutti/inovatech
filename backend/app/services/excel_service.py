from io import BytesIO
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.core.security import criar_hash_senha
from app.models.consulta import Consulta
from app.models.lancamento_financeiro import LancamentoFinanceiro
from app.models.log_auditoria import LogAuditoria
from app.models.medico import Medico
from app.models.paciente import Paciente
from app.models.prontuario import Prontuario
from app.models.registro_ponto import RegistroPonto
from app.models.usuario import Usuario
from app.repositories import (
    consulta_repository,
    financeiro_repository,
    medico_repository,
    paciente_repository,
    ponto_repository,
    prontuario_repository,
    usuario_repository,
)
from app.services.ponto_service import calcular_dados_ponto


def exportar_para_xlsx(dados: list[dict], colunas: list[str]) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, coluna in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, registro in enumerate(dados, 2):
        for col_idx, coluna in enumerate(colunas, 1):
            valor = registro.get(coluna, "")
            ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor is not None else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_pacientes(db: Session) -> BytesIO:
    pacientes = db.query(Paciente).all()
    dados = [
        {
            "id": p.id,
            "nome_completo": p.pessoa.nome_completo,
            "cpf": p.pessoa.cpf,
            "telefone": p.pessoa.telefone,
            "email": p.pessoa.email,
            "status": p.pessoa.status,
            "data_nascimento": str(p.data_nascimento or ""),
            "convenio": p.convenio,
            "endereco": p.endereco,
        }
        for p in pacientes
    ]
    colunas = ["id", "nome_completo", "cpf", "telefone", "email", "status", "data_nascimento", "convenio", "endereco"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_medicos(db: Session) -> BytesIO:
    medicos = db.query(Medico).all()
    dados = [
        {
            "id": m.id,
            "nome_completo": m.pessoa.nome_completo,
            "cpf": m.pessoa.cpf,
            "crm": m.crm,
            "especialidade": m.especialidade,
            "telefone": m.pessoa.telefone,
            "email": m.pessoa.email,
            "status": m.pessoa.status,
            "data_formatura": str(m.data_formatura or ""),
        }
        for m in medicos
    ]
    colunas = ["id", "nome_completo", "cpf", "crm", "especialidade", "telefone", "email", "status", "data_formatura"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_consultas(db: Session) -> BytesIO:
    consultas = db.query(Consulta).all()
    dados = [
        {
            "id": c.id,
            "paciente_id": c.paciente_id,
            "medico_id": c.medico_id,
            "data": str(c.data),
            "horario": str(c.horario),
            "tipo_consulta": c.tipo_consulta,
            "convenio": c.convenio,
            "valor": float(c.valor),
            "status": c.status,
        }
        for c in consultas
    ]
    colunas = ["id", "paciente_id", "medico_id", "data", "horario", "tipo_consulta", "convenio", "valor", "status"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_prontuarios(db: Session) -> BytesIO:
    prontuarios = db.query(Prontuario).all()
    dados = [
        {
            "id": p.id,
            "paciente_id": p.paciente_id,
            "medico_id": p.medico_id,
            "data": str(p.data),
            "cid": p.cid,
            "diagnostico": p.diagnostico,
            "prescricao": p.prescricao,
            "retorno_em_dias": p.retorno_em_dias,
            "laudo_liberado": p.laudo_liberado,
        }
        for p in prontuarios
    ]
    colunas = ["id", "paciente_id", "medico_id", "data", "cid", "diagnostico", "prescricao", "retorno_em_dias", "laudo_liberado"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_financeiro(db: Session) -> BytesIO:
    lancamentos = db.query(LancamentoFinanceiro).all()
    dados = [
        {
            "id": l.id,
            "paciente_id": l.paciente_id,
            "medico_id": l.medico_id,
            "consulta_id": l.consulta_id,
            "data": str(l.data),
            "servico": l.servico,
            "convenio": l.convenio,
            "valor": float(l.valor),
            "status": l.status,
            "forma_pagamento": l.forma_pagamento,
            "observacao": l.observacao,
        }
        for l in lancamentos
    ]
    colunas = ["id", "paciente_id", "medico_id", "consulta_id", "data", "servico", "convenio", "valor", "status", "forma_pagamento", "observacao"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_ponto(db: Session) -> BytesIO:
    pontos = db.query(RegistroPonto).all()
    dados = [
        {
            "id": p.id,
            "usuario_id": p.usuario_id,
            "data": str(p.data),
            "entrada": str(p.entrada or ""),
            "saida": str(p.saida or ""),
            "h_trabalhadas": float(p.h_trabalhadas or 0),
            "h_esperadas": float(p.h_esperadas),
            "diferenca": float(p.diferenca or 0),
            "situacao": p.situacao,
        }
        for p in pontos
    ]
    colunas = ["id", "usuario_id", "data", "entrada", "saida", "h_trabalhadas", "h_esperadas", "diferenca", "situacao"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_usuarios(db: Session) -> BytesIO:
    usuarios = db.query(Usuario).all()
    dados = [
        {
            "id": u.id,
            "nome": u.nome,
            "perfil": u.perfil,
            "login": u.login,
            "email": u.email,
            "status": u.status,
            "ultimo_acesso": str(u.ultimo_acesso or ""),
            "modulos_permitidos": ",".join(u.modulos_permitidos or []),
            "observacao": u.observacao,
        }
        for u in usuarios
    ]
    colunas = ["id", "nome", "perfil", "login", "email", "status", "ultimo_acesso", "modulos_permitidos", "observacao"]
    return exportar_para_xlsx(dados, colunas), len(dados)


def exportar_log_auditoria(db: Session) -> BytesIO:
    logs = db.query(LogAuditoria).order_by(LogAuditoria.data_hora.desc()).all()
    dados = [
        {
            "id": l.id,
            "data_hora": str(l.data_hora),
            "usuario_id": l.usuario_id,
            "acao": l.acao,
            "modulo": l.modulo,
            "ip": l.ip,
            "resultado": l.resultado,
            "detalhes": l.detalhes,
        }
        for l in logs
    ]
    colunas = ["id", "data_hora", "usuario_id", "acao", "modulo", "ip", "resultado", "detalhes"]
    return exportar_para_xlsx(dados, colunas), len(dados)


async def importar_de_xlsx(
    arquivo: UploadFile, colunas_obrigatorias: list[str]
) -> tuple[list[dict], list[str]]:
    conteudo = await arquivo.read()
    try:
        df = pd.read_excel(BytesIO(conteudo), dtype=str)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erro ao ler arquivo: {str(e)}")

    colunas_faltando = [c for c in colunas_obrigatorias if c not in df.columns]
    if colunas_faltando:
        raise HTTPException(
            status_code=422,
            detail=f"Colunas obrigatórias ausentes: {', '.join(colunas_faltando)}",
        )

    registros_validos = []
    erros = []

    for idx, row in df.iterrows():
        linha_num = idx + 2
        try:
            faltando = [c for c in colunas_obrigatorias if pd.isna(row.get(c))]
            if faltando:
                erros.append(f"Linha {linha_num}: campos obrigatórios vazios: {', '.join(faltando)}")
                continue
            registro = {col: (None if pd.isna(val) else val) for col, val in row.items()}
            registros_validos.append(registro)
        except Exception as e:
            erros.append(f"Linha {linha_num}: {str(e)}")

    return registros_validos, erros


def importar_pacientes(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            cpf = _texto_obrigatorio(registro, "cpf")
            if paciente_repository.buscar_por_cpf(db, cpf):
                erros.append(f"Linha {linha}: CPF já cadastrado")
                continue
            pessoa_dados = {
                "nome_completo": _texto_obrigatorio(registro, "nome_completo"),
                "cpf": cpf,
                "telefone": _texto(registro, "telefone"),
                "email": _texto(registro, "email"),
            }
            status = _texto(registro, "status")
            if status:
                pessoa_dados["status"] = _enum(status, {"ativo", "inativo"}, "status")
            paciente_dados = {
                "data_nascimento": _data(registro, "data_nascimento"),
                "convenio": _texto(registro, "convenio"),
                "endereco": _texto(registro, "endereco"),
            }
            paciente_repository.criar(db, pessoa_dados, paciente_dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_medicos(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            cpf = _texto_obrigatorio(registro, "cpf")
            crm = _texto_obrigatorio(registro, "crm")
            if medico_repository.buscar_por_cpf(db, cpf):
                erros.append(f"Linha {linha}: CPF já cadastrado")
                continue
            if medico_repository.buscar_por_crm(db, crm):
                erros.append(f"Linha {linha}: CRM já cadastrado")
                continue
            pessoa_dados = {
                "nome_completo": _texto_obrigatorio(registro, "nome_completo"),
                "cpf": cpf,
                "telefone": _texto(registro, "telefone"),
                "email": _texto(registro, "email"),
            }
            status = _texto(registro, "status")
            if status:
                pessoa_dados["status"] = _enum(status, {"ativo", "inativo"}, "status")
            medico_dados = {
                "crm": crm,
                "especialidade": _texto_obrigatorio(registro, "especialidade"),
                "data_formatura": _data(registro, "data_formatura"),
            }
            medico_repository.criar(db, pessoa_dados, medico_dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_consultas(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            paciente_id = _inteiro_obrigatorio(registro, "paciente_id")
            medico_id = _inteiro_obrigatorio(registro, "medico_id")
            _existe(db, Paciente, paciente_id, "paciente_id")
            _existe(db, Medico, medico_id, "medico_id")
            dados = {
                "paciente_id": paciente_id,
                "medico_id": medico_id,
                "data": _data_obrigatoria(registro, "data"),
                "horario": _hora_obrigatoria(registro, "horario"),
                "tipo_consulta": _texto_obrigatorio(registro, "tipo_consulta"),
                "convenio": _texto(registro, "convenio"),
                "valor": _decimal_obrigatorio(registro, "valor"),
                "status": _enum(_texto(registro, "status") or "agendada", {"agendada", "confirmada", "realizada", "cancelada"}, "status"),
            }
            consulta_repository.criar(db, dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_prontuarios(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            paciente_id = _inteiro_obrigatorio(registro, "paciente_id")
            medico_id = _inteiro_obrigatorio(registro, "medico_id")
            _existe(db, Paciente, paciente_id, "paciente_id")
            _existe(db, Medico, medico_id, "medico_id")
            dados = {
                "paciente_id": paciente_id,
                "medico_id": medico_id,
                "data": _data_obrigatoria(registro, "data"),
                "cid": _texto_obrigatorio(registro, "cid"),
                "diagnostico": _texto_obrigatorio(registro, "diagnostico"),
                "prescricao": _texto_obrigatorio(registro, "prescricao"),
                "retorno_em_dias": _inteiro(registro, "retorno_em_dias") or 0,
                "laudo_liberado": _booleano(registro, "laudo_liberado") or False,
            }
            prontuario_repository.criar(db, dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_financeiro(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            paciente_id = _inteiro_obrigatorio(registro, "paciente_id")
            _existe(db, Paciente, paciente_id, "paciente_id")
            dados = {
                "consulta_id": _inteiro(registro, "consulta_id"),
                "paciente_id": paciente_id,
                "medico_id": _inteiro(registro, "medico_id"),
                "data": _data_obrigatoria(registro, "data"),
                "servico": _texto_obrigatorio(registro, "servico"),
                "convenio": _texto(registro, "convenio"),
                "valor": _decimal_obrigatorio(registro, "valor"),
                "status": _enum(_texto(registro, "status") or "pendente", {"pago", "pendente", "atrasado"}, "status"),
                "forma_pagamento": _texto(registro, "forma_pagamento"),
                "observacao": _texto(registro, "observacao"),
            }
            if dados["medico_id"]:
                _existe(db, Medico, dados["medico_id"], "medico_id")
            financeiro_repository.criar(db, dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_ponto(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            usuario_id = _inteiro_obrigatorio(registro, "usuario_id")
            _existe(db, Usuario, usuario_id, "usuario_id")
            dados = {
                "usuario_id": usuario_id,
                "data": _data_obrigatoria(registro, "data"),
                "entrada": _hora(registro, "entrada"),
                "saida": _hora(registro, "saida"),
                "h_esperadas": _decimal(registro, "h_esperadas") or Decimal("8.0"),
            }
            _calcular_ponto(dados)
            ponto_repository.criar(db, dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def importar_usuarios(db: Session, registros: list[dict]) -> tuple[int, list[str]]:
    importados = 0
    erros: list[str] = []
    for linha, registro in _linhas(registros):
        try:
            login = _texto_obrigatorio(registro, "login")
            email = _texto_obrigatorio(registro, "email")
            if usuario_repository.buscar_por_login(db, login):
                erros.append(f"Linha {linha}: login já cadastrado")
                continue
            if usuario_repository.buscar_por_email(db, email):
                erros.append(f"Linha {linha}: email já cadastrado")
                continue
            modulos = _texto(registro, "modulos_permitidos")
            dados = {
                "nome": _texto_obrigatorio(registro, "nome"),
                "perfil": _enum(_texto_obrigatorio(registro, "perfil"), {"gestor", "recepcionista", "medico", "paciente"}, "perfil"),
                "login": login,
                "email": email,
                "password_hash": criar_hash_senha(_texto(registro, "password") or "Inovatech@2026"),
                "status": _enum(_texto(registro, "status") or "ativo", {"ativo", "inativo"}, "status"),
                "modulos_permitidos": [m.strip() for m in modulos.split(",")] if modulos else None,
                "observacao": _texto(registro, "observacao"),
            }
            usuario_repository.criar(db, dados)
            importados += 1
        except Exception as exc:
            erros.append(f"Linha {linha}: {exc}")
    return importados, erros


def _linhas(registros: list[dict]):
    for idx, registro in enumerate(registros, start=2):
        yield idx, registro


def _valor(registro: dict, campo: str) -> Any:
    valor = registro.get(campo)
    if valor is None or pd.isna(valor):
        return None
    if isinstance(valor, str):
        valor = valor.strip()
        return valor or None
    return valor


def _texto(registro: dict, campo: str) -> str | None:
    valor = _valor(registro, campo)
    return None if valor is None else str(valor).strip()


def _texto_obrigatorio(registro: dict, campo: str) -> str:
    valor = _texto(registro, campo)
    if not valor:
        raise ValueError(f"{campo} obrigatorio")
    return valor


def _inteiro(registro: dict, campo: str) -> int | None:
    valor = _valor(registro, campo)
    if valor is None:
        return None
    return int(float(str(valor).replace(",", ".")))


def _inteiro_obrigatorio(registro: dict, campo: str) -> int:
    valor = _inteiro(registro, campo)
    if valor is None:
        raise ValueError(f"{campo} obrigatorio")
    return valor


def _decimal(registro: dict, campo: str) -> Decimal | None:
    valor = _valor(registro, campo)
    if valor is None:
        return None
    return Decimal(str(valor).replace(",", "."))


def _decimal_obrigatorio(registro: dict, campo: str) -> Decimal:
    valor = _decimal(registro, campo)
    if valor is None:
        raise ValueError(f"{campo} obrigatorio")
    return valor


def _data(registro: dict, campo: str) -> date | None:
    valor = _valor(registro, campo)
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return pd.to_datetime(valor).date()


def _data_obrigatoria(registro: dict, campo: str) -> date:
    valor = _data(registro, campo)
    if valor is None:
        raise ValueError(f"{campo} obrigatorio")
    return valor


def _hora(registro: dict, campo: str) -> time | None:
    valor = _valor(registro, campo)
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    texto = str(valor)
    if len(texto.split(":")) == 2:
        texto = f"{texto}:00"
    return time.fromisoformat(texto)


def _hora_obrigatoria(registro: dict, campo: str) -> time:
    valor = _hora(registro, campo)
    if valor is None:
        raise ValueError(f"{campo} obrigatorio")
    return valor


def _booleano(registro: dict, campo: str) -> bool | None:
    valor = _valor(registro, campo)
    if valor is None:
        return None
    if isinstance(valor, bool):
        return valor
    texto = str(valor).strip().lower()
    if texto in {"true", "1", "sim", "s", "yes"}:
        return True
    if texto in {"false", "0", "nao", "não", "n", "no"}:
        return False
    raise ValueError(f"{campo} invalido")


def _enum(valor: str, permitidos: set[str], campo: str) -> str:
    valor_normalizado = valor.strip().lower()
    if valor_normalizado not in permitidos:
        raise ValueError(f"{campo} inválido")
    return valor_normalizado


def _existe(db: Session, modelo, item_id: int, campo: str) -> None:
    if not db.get(modelo, item_id):
        raise ValueError(f"{campo} não encontrado")


def _calcular_ponto(dados: dict) -> None:
    resultado = calcular_dados_ponto(
        dados.get("entrada"),
        dados.get("saida"),
        Decimal(str(dados.get("h_esperadas") or "8.0")),
    )
    if resultado is not None:
        dados["h_trabalhadas"], dados["diferenca"], dados["situacao"] = resultado
